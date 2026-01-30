"""
Export selected pd.* tables to a single Excel workbook.

Outputs: report/data_export.xlsx
Sheets: one per table + summary_counts

Dependencies: openpyxl, psycopg v3, python-dotenv
Beginner-readable: no ORM, clear comments.
"""
from __future__ import annotations

from typing import Sequence
from pathlib import Path

from openpyxl import Workbook

from .db import connect


TABLES: Sequence[str] = (
    "accounts",
    "contacts",
    "opportunities",
    "opportunity_sponsorships",
    "applications",
    "experiences",
    "work_experiences",
    "partner_engagements",
    "outreach_messages",
    "account_aliases",
    "etl_school_name_review_queue",
    "entity_embeddings",
    "query_audit_log",
)


def fetch_all_rows(table: str):
    """Return (columns, rows) for pd.<table>."""
    with connect() as conn, conn.cursor() as cur:
        cur.execute(f"select * from pd.{table}")
        cols = [desc[0] for desc in cur.description]
        rows = cur.fetchall()
        return cols, rows


def write_sheet(wb: Workbook, name: str, columns: list[str], rows: list[tuple]):
    """Create a sheet and write header + data rows."""
    ws = wb.create_sheet(title=name)
    ws.append(columns)
    for r in rows:
        ws.append(list(r))


def main() -> Path:
    """Export all TABLES into an Excel workbook and return path."""
    out_path = Path(__file__).resolve().parents[1] / "report" / "data_export.xlsx"
    wb = Workbook()
    # Remove default sheet created by openpyxl
    default = wb.active
    wb.remove(default)

    summary_rows = []
    for t in TABLES:
        try:
            cols, rows = fetch_all_rows(t)
            write_sheet(wb, t, cols, rows)
            summary_rows.append((t, len(rows)))
        except Exception as e:
            # Create a sheet with the error message so the user sees it
            ws = wb.create_sheet(title=f"{t}_error")
            ws.append(["error"])
            ws.append([str(e)])
            summary_rows.append((t, f"ERROR: {e}"))

    # Add summary sheet
    ws_sum = wb.create_sheet(title="summary_counts")
    ws_sum.append(["table_name", "count"])
    for name, count in summary_rows:
        ws_sum.append([name, count])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = main()
    print(f"✅ Exported Excel to: {path}")
